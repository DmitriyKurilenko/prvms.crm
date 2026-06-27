"""Импорт/экспорт контактов и компаний.

Изолирует stdlib `csv` и внешний `openpyxl` (контракт сверен с установленной
версией 3.1.5: `load_workbook(file, read_only=True, data_only=True)` принимает
file-like, `ws.iter_rows(values_only=True)` отдаёт кортежи значений; в read-only
режиме воркбук нужно закрывать). Экспорт повторяет паттерн `apps/audit/api.py`
(StringIO → csv.writer → text/csv), добавляя BOM `﻿` для корректного
открытия кириллицы в Excel.
"""
from __future__ import annotations

import csv
import io

# Целевые поля модели, доступные для маппинга при импорте.
# Колонки, отображённые на ключ не из этого набора, уходят в custom_fields.
CONTACT_FIELDS: tuple[str, ...] = (
    'first_name', 'last_name', 'phone', 'email', 'position', 'messenger_id', 'source',
)
COMPANY_FIELDS: tuple[str, ...] = (
    'name', 'inn', 'phone', 'email', 'address', 'website',
)
# Для сделок `pipeline`/`stage`/`contact` — не прямые поля модели, а ссылки,
# которые разрешаются по имени/телефону при импорте (см. tasks._import_one_deal).
DEAL_FIELDS: tuple[str, ...] = (
    'name', 'amount', 'currency', 'source', 'pipeline', 'stage', 'contact',
)

# Подсказки авто-маппинга: нормализованный заголовок колонки → поле модели.
_CONTACT_HEADER_HINTS: dict[str, str] = {
    'имя': 'first_name', 'name': 'first_name', 'first_name': 'first_name', 'firstname': 'first_name',
    'фамилия': 'last_name', 'last_name': 'last_name', 'lastname': 'last_name', 'surname': 'last_name',
    'телефон': 'phone', 'phone': 'phone', 'tel': 'phone', 'мобильный': 'phone',
    'email': 'email', 'почта': 'email', 'e-mail': 'email', 'mail': 'email',
    'должность': 'position', 'position': 'position',
    'мессенджер': 'messenger_id', 'messenger': 'messenger_id', 'messenger_id': 'messenger_id',
    'источник': 'source', 'source': 'source',
}
_COMPANY_HEADER_HINTS: dict[str, str] = {
    'название': 'name', 'наименование': 'name', 'компания': 'name', 'name': 'name', 'company': 'name',
    'инн': 'inn', 'inn': 'inn',
    'телефон': 'phone', 'phone': 'phone',
    'email': 'email', 'почта': 'email', 'e-mail': 'email',
    'адрес': 'address', 'address': 'address',
    'сайт': 'website', 'website': 'website', 'url': 'website',
}
_DEAL_HEADER_HINTS: dict[str, str] = {
    'название': 'name', 'сделка': 'name', 'name': 'name', 'title': 'name',
    'сумма': 'amount', 'amount': 'amount', 'цена': 'amount',
    'валюта': 'currency', 'currency': 'currency',
    'источник': 'source', 'source': 'source',
    'воронка': 'pipeline', 'pipeline': 'pipeline',
    'стадия': 'stage', 'этап': 'stage', 'stage': 'stage',
    'контакт': 'contact', 'contact': 'contact', 'клиент': 'contact',
}


def parse_file(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    """Возвращает (заголовки, строки-словари column→value). Поддержка .csv и .xlsx."""
    lower = (filename or '').lower()
    if lower.endswith('.csv'):
        text = content.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = [{h: (row.get(h) or '') for h in headers} for row in reader]
        return headers, rows
    if lower.endswith('.xlsx'):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                first = next(rows_iter)
            except StopIteration:
                return [], []
            headers = [str(h).strip() if h is not None else '' for h in first]
            rows: list[dict] = []
            for raw in rows_iter:
                values = ['' if v is None else str(v) for v in raw]
                # выравниваем длину строки под заголовки
                values += [''] * (len(headers) - len(values))
                rows.append(dict(zip(headers, values, strict=False)))
            return headers, rows
        finally:
            wb.close()
    raise ValueError('Поддерживаются только файлы .csv и .xlsx')


def suggest_mapping(entity: str, headers: list[str]) -> dict[str, str]:
    """Авто-сопоставление заголовков колонок целевым полям модели (по словарю подсказок)."""
    hints = {
        'contacts': _CONTACT_HEADER_HINTS,
        'companies': _COMPANY_HEADER_HINTS,
        'deals': _DEAL_HEADER_HINTS,
    }.get(entity, _CONTACT_HEADER_HINTS)
    mapping: dict[str, str] = {}
    for header in headers:
        key = str(header or '').strip().lower()
        if key in hints:
            mapping[header] = hints[key]
    return mapping


def allowed_fields(entity: str) -> tuple[str, ...]:
    return {
        'contacts': CONTACT_FIELDS,
        'companies': COMPANY_FIELDS,
        'deals': DEAL_FIELDS,
    }.get(entity, CONTACT_FIELDS)


def _csv_response_body(header_row: list[str], data_rows) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header_row)
    for row in data_rows:
        writer.writerow(row)
    # BOM: Excel распознаёт кодировку UTF-8 и не ломает кириллицу.
    return '﻿' + buf.getvalue()


def export_contacts_csv(queryset) -> str:
    rows = (
        [c.id, c.first_name, c.last_name, c.phone, c.email,
         c.company.name if c.company else '', c.position, c.source,
         c.created_at.isoformat()]
        for c in queryset
    )
    return _csv_response_body(
        ['id', 'first_name', 'last_name', 'phone', 'email', 'company', 'position', 'source', 'created_at'],
        rows,
    )


def export_companies_csv(queryset) -> str:
    rows = (
        [c.id, c.name, c.inn, c.phone, c.email, c.address, c.website, c.created_at.isoformat()]
        for c in queryset
    )
    return _csv_response_body(
        ['id', 'name', 'inn', 'phone', 'email', 'address', 'website', 'created_at'],
        rows,
    )


def export_deals_csv(queryset) -> str:
    def _contact_label(d) -> str:
        if not d.contact:
            return ''
        return f'{d.contact.first_name} {d.contact.last_name}'.strip() or d.contact.phone

    rows = (
        [d.id, d.name, d.pipeline.name if d.pipeline else '', d.stage.name if d.stage else '',
         float(d.amount) if d.amount is not None else '', d.currency, _contact_label(d),
         d.source, d.created_at.isoformat()]
        for d in queryset
    )
    return _csv_response_body(
        ['id', 'name', 'pipeline', 'stage', 'amount', 'currency', 'contact', 'source', 'created_at'],
        rows,
    )
